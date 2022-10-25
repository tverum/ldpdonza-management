from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Lid, Ouder, Functie, PloegLid, Seizoen

# A mapping of the string representation of Lid to the actual field.
# No better reverse determination of fields was found so bit of a hack
LID_FIELDS = dict(
    zip(
        [str(field) for field in Lid._meta.get_fields() if not field.is_relation],
        [field for field in Lid._meta.get_fields() if not field.is_relation],
    )
)

OUDER_FIELDS = ["ouder 1: gsm", "ouder 1: email", "ouder 2: gsm", "ouder 2: email"]
BETALING_FIELDS = ["origineel bedrag", "afgelost bedrag"]


class CoachLidDownloadResource(resources.ModelResource):
    ouder_1_gsm = fields.Field(
        attribute="moeder",
        column_name="Ouder 1: Gsmnummer",
        widget=ForeignKeyWidget(Ouder, "gsmnummer"),
    )
    ouder_2_gsm = fields.Field(
        attribute="vader",
        column_name="Ouder 2: Gsmnummer",
        widget=ForeignKeyWidget(Ouder, "gsmnummer"),
    )
    ouder_1_mail = fields.Field(
        attribute="moeder",
        column_name="Ouder 1: Mail",
        widget=ForeignKeyWidget(Ouder, "email"),
    )
    ouder_2_mail = fields.Field(
        attribute="vader",
        column_name="Ouder 2: Mail",
        widget=ForeignKeyWidget(Ouder, "email"),
    )

    class Meta:
        model = Lid
        fields = (
            "voornaam",
            "familienaam",
            "gsmnummer",
            "email",
            "ouder_1_gsm",
            "ouder_2_gsm",
            "ouder_1_mail",
            "ouder_2_mail",
        )


def create_team_workbook(queryset_coaches, queryset_spelers, queryset_pvn, queryset_hh):
    workbook = Workbook()

    # --- SPELERS ---
    # Get active worksheet/tab
    columns = [
        "Voornaam",
        "Familienaam",
        "Gsmnummer",
        "Email",
        "GSM Ouder 1",
        "Email Ouder 1",
        "GSM Ouder 2",
        "Email Ouder 2",
    ]
    worksheet = workbook.active
    worksheet.title = "Spelers"

    cell = worksheet.cell(row=2, column=2)
    cell.value = "SPELERS"
    cell.font = Font(name="Calibri", bold=True, size=20)
    row_num = 4

    create_team_sheet(columns, queryset_spelers, row_num, worksheet, minimal=False)

    # --- COACHES ---
    # Create new worksheet
    columns = ["Voornaam", "Familienaam", "Gsmnummer", "Email"]
    worksheet = workbook.create_sheet(
        title="Coaches",
        index=1,
    )

    cell = worksheet.cell(row=2, column=2)
    cell.value = "COACHES"
    cell.font = Font(name="Calibri", bold=True, size=20)
    row_num = 4

    create_team_sheet(columns, queryset_coaches, row_num, worksheet, minimal=True)

    # --- PLOEGVERANTWOORDELIJKEN ---
    columns = ["Voornaam", "Familienaam", "Gsmnummer", "Email"]
    worksheet = workbook.create_sheet(
        title="Ploegverantwoordelijken",
        index=1,
    )

    cell = worksheet.cell(row=2, column=2)
    cell.value = "PLOEGVERANTWOORDELIJKEN"
    cell.font = Font(name="Calibri", bold=True, size=20)
    row_num = 4

    create_team_sheet(columns, queryset_pvn, row_num, worksheet, minimal=True)

    # --- HELPENDE HANDEN ---
    columns = ["Voornaam", "Familienaam", "Gsmnummer", "Email"]
    worksheet = workbook.create_sheet(
        title="Helpende Handen",
        index=1,
    )

    cell = worksheet.cell(row=2, column=2)
    cell.value = "HELPENDE HANDEN"
    cell.font = Font(name="Calibri", bold=True, size=20)
    row_num = 4

    create_team_sheet(columns, queryset_hh, row_num, worksheet, minimal=True)

    return workbook


def create_team_sheet(columns, queryset, row_num, worksheet, minimal):
    """
    Create a sheet with a specific queryset
    :param columns: the column names
    :param queryset: the queryset to put into the excel sheet
    :param row_num: the row number to start from
    :param worksheet: the worksheet to write the data to
    :param minimal: minimal data writing (for coaches and pv)
    """
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 2):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all movies
    for lid in queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            lid.voornaam,
            lid.familienaam,
            str(lid.gsmnummer),
            lid.email,
        ]

        if not minimal:
            if lid.moeder:
                row.append(str(lid.moeder.gsmnummer))
                row.append(lid.moeder.email)
            else:
                row.append("")
                row.append("")

            if lid.vader:
                row.append(str(lid.vader.gsmnummer))
                row.append(lid.vader.email)
            else:
                row.append("")
                row.append("")

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


def create_general_workbook(ploegen, selected_fields):
    workbook = Workbook()

    include_ouders = False
    include_betaling = False

    seizoen = ploegen[0].seizoen

    if "management.ouders" in selected_fields:
        selected_fields = [
            field for field in selected_fields if field != "management.ouders"
        ]
        include_ouders = True
    if "management.betaling" in selected_fields:
        selected_fields = [
            field for field in selected_fields if field != "management.betaling"
        ]
        include_betaling = True

    selected_fields = [LID_FIELDS[field] for field in selected_fields]

    # Get active worksheet/tab
    worksheet = workbook.active

    # --- SPELERS ---
    worksheet.title = "Spelers"

    speler = Functie.objects.get(functie="Speler")
    create_sheet(
        ploegen,
        selected_fields,
        worksheet,
        speler,
        include_ouders,
        include_betaling,
        seizoen,
    )

    # --- COACHES ---
    worksheet = workbook.create_sheet(
        title="Coaches",
        index=1,
    )
    coach = Functie.objects.get(functie="Coach")
    create_sheet(ploegen, selected_fields, worksheet, coach, False, False, seizoen)

    # --- PLOEGVERANTWOORDELIJKEN ---
    worksheet = workbook.create_sheet(
        title="Ploegverantwoordelijken",
        index=1,
    )
    ploegverantwoordelijke = Functie.objects.get(functie="Ploegverantwoordelijke")
    create_sheet(
        ploegen,
        selected_fields,
        worksheet,
        ploegverantwoordelijke,
        False,
        False,
        seizoen,
    )

    # --- HELPENDE HANDEN ---
    worksheet = workbook.create_sheet(
        title="Helpende handen",
        index=1,
    )
    helpende_handen = Functie.objects.get(functie="Helpende Handen")
    create_sheet(
        ploegen, selected_fields, worksheet, helpende_handen, False, False, seizoen
    )
    return workbook


def create_sheet(
    ploegen,
    selected_fields,
    worksheet,
    functie: Functie,
    include_ouders: bool,
    include_betaling: bool,
    seizoen: Seizoen,
):
    """
    Create a sheet for the specific function
    :param ploegen: the ploegen for which the exported xlsx should be generated
    :param selected_fields: the fields that should be shown
    :param worksheet:
    :param functie: the functie for which to create the sheet
    :param include_ouders: if the export should include the ouders or not
    :param include_betaling: if the export should include payment records or not
    :param seizoen: the seizoen to export the betalingen for
    :return:
    """
    cell = worksheet.cell(row=2, column=2)
    cell.value = functie.functie
    cell.font = Font(name="Calibri", bold=True, size=20)
    row_num = 4
    # Assign the titles for each cell of the header
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = "Ploeg"

    print(selected_fields)

    col_num = 0
    for col_num, field in enumerate(selected_fields, 2):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = field.verbose_name
    if include_ouders:
        col_num += 1
        offset = 0
        for offset, field in enumerate(OUDER_FIELDS):
            cell = worksheet.cell(row=row_num, column=col_num + offset)
            cell.value = field
        col_num += offset
    if include_betaling:
        col_num += 1
        for offset, field in enumerate(BETALING_FIELDS):
            cell = worksheet.cell(row=row_num, column=col_num + offset)
            cell.value = field

    for ploeg in ploegen:
        ploegleden = PloegLid.objects.filter(ploeg=ploeg, functie=functie)
        lid_ids = [ploeglid.lid.club_id for ploeglid in ploegleden]
        queryset_spelers = Lid.objects.filter(club_id__in=lid_ids)

        for lid in queryset_spelers:
            row_num += 1

            if "management.ouders" in selected_fields:
                s_fields = [
                    field
                    for field in selected_fields
                    if field is not "management.ouders"
                ]
                pass

            row = [
                str(getattr(lid, field.name))
                if getattr(lid, field.name) is not None
                else ""
                for field in selected_fields
            ]

            if include_ouders:
                if lid.vader:
                    row.append(str(lid.vader.gsmnummer))
                    row.append(str(lid.vader.email))
                else:
                    row.append("")
                    row.append("")
                if lid.moeder:
                    row.append(str(lid.moeder.gsmnummer))
                    row.append(str(lid.moeder.email))
                else:
                    row.append("")
                    row.append("")
            if include_betaling:
                betaling = lid.betaling_set.filter(seizoen=seizoen)
                if betaling:
                    row.append(betaling[0].origineel_bedrag)
                    row.append(betaling[0].afgelost_bedrag)
                else:
                    row.append("nvt.")
                    row.append("nvt.")

            # Assign the data for each cell of the row
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = ploeg.naam

            for col_num, cell_value in enumerate(row, 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
